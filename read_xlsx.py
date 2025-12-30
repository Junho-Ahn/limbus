import openpyxl
import json

# xlsx 파일 읽기
wb = openpyxl.load_workbook('엔케팔린 효율 계산기 v1.1.2.xlsx', data_only=True)

# 모든 시트 데이터를 딕셔너리로 저장
data = {}

def convert_value(value):
    """값을 JSON 직렬화 가능한 형태로 변환"""
    if value is None:
        return None
    if isinstance(value, (int, float, str, bool)):
        return value
    # 다른 타입은 문자열로 변환
    return str(value)

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    sheet_data = []
    
    for row in sheet.iter_rows(values_only=True):
        # None이 아닌 값이 있는 행만 추가
        row_values = [convert_value(cell) for cell in row]
        if any(cell is not None for cell in row_values):
            sheet_data.append(row_values)
    
    data[sheet_name] = sheet_data

# JSON으로 출력
print(json.dumps(data, ensure_ascii=False, indent=2))


